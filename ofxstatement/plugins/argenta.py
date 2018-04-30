# -*- coding: utf-8 -*-

import re
import itertools
import logging

from datetime import datetime
from openpyxl import load_workbook

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine, BankAccount

def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(itertools.islice(iterable, n))

class ArgentaStatementParser(StatementParser):
    header = ['Rekening', 'Boekdatum', 'Valutadatum', 'Referentie',
    'Beschrijving', 'Bedrag', 'Munt', 'Verrichtingsdatum',
    'Rekening tegenpartij', 'Naam tegenpartij', 'Mededeling']
    dict_transaction_types = {
        'Inkomende overschrijving': 'CREDIT',
        'Uitgaande overschrijving': 'DEBIT',
        'Betaling Bancontact': 'POS',
        'Betaling Maestro': 'POS',
        'Bestendige opdracht': 'REPEATPMT',
        'SEPA-domiciliëring': 'DIRECTDEBIT'
    }# TODO Add more transaction types as I encounter them.

    def __init__(self, fin):
        """Create a new ArgentaStatementParser instance.
        """
        
        self.col_index = dict(zip(self.header, range(0, 11)))

        self.sheet = load_workbook(filename=fin, read_only=True).active
        # dimensions are incorrectly set in the XLSX-file
        self.sheet.max_row = self.sheet.max_column = None
        
        self.validate()

        self.statement = self.parse_statement()

    def validate(self):
        """
        Naive validation to make sure that the XLSX document is structured
        the way it was when this parser was written.

        :raises ValueError if workbook has invalid format
        """

        try:
            self._validate()
        except AssertionError as e:
            raise ValueError(e)

    def _validate(self):
        """
        Headers of the XLSX file:
        [Rekening, Boekdatum, Valutadatum, Referentie,
         Beschrijving, Bedrag, Munt, Verrichtingsdatum,
         Rekening tegenpartij, Naam tegenpartij, Mededeling]
        Row of sample data:
        [BE46 5390 0754 7034, 24-04-2018, 24-04-2018, 7SLLCW65XDZQU891,
         Uitgaande overschrijving, -125,00, EUR, 24-04-2018,
         BE68 5390 0754 7034, Adams White, Comment of transaction]
        """
        
        logging.info('Verifying that the sheet has at least 2 rows.')
        top_two_rows = take(2, self.sheet.iter_rows())
        assert len(top_two_rows) == 2

        logging.info('Verifying that the first row has 11 cells.')
        assert len(top_two_rows[1]) == 11

        logging.info('Verifying statements header.')
        statement_header_row = [c.value for c in top_two_rows[0]]
        assert self.header == statement_header_row

        logging.info('Verifying account numbers are IBAN Belgian formatted (A2 and I2).')
        first_stmt_row = [c.value for c in top_two_rows[1]]
        assert BankAccountIban.is_valid(first_stmt_row[self.col_index['Rekening']])
        assert BankAccountIban.is_valid(first_stmt_row[self.col_index['Rekening tegenpartij']])
        
        logging.info('Verifying statement date is a date (H2).')
        assert isinstance(first_stmt_row[self.col_index['Verrichtingsdatum']], datetime)
        
        logging.info('Verifying account id is equal for every transaction (column A).')
        BankAccountIbans_are_equal = True
        currencies_are_equal = True
        for row in self.sheet.iter_rows(min_row=3):
            # Rekening A
            if row[0].value != first_stmt_row[self.col_index['Rekening']]:
                BankAccountIbans_are_equal = False
                break
            # Munt G
            if row[6].value != first_stmt_row[self.col_index['Munt']]:
                currencies_are_equal = False
                break
        
        assert BankAccountIbans_are_equal
        logging.info('Verifying currency is equal for every transaction (column G).')
        assert currencies_are_equal

        logging.info('Everything is OK!')

    def parse_statement(self):
        statement = Statement()
        
        account_holder = BankAccountIban(self.sheet['A2'].value)
        statement.bank_id = account_holder.bank_id
        statement.account_id = account_holder.acct_id
        statement.currency = self.sheet['G2'].value
        
        return statement

    def split_records(self):
        # Skip the header
        for row in self.sheet.iter_rows(min_row=2):
            yield [c.value for c in row]

    def parse_record(self, row):
        stmt_line = StatementLine()
        
        stmt_line.id = row[self.col_index['Referentie']]
        stmt_line.date = row[self.col_index['Verrichtingsdatum']]
        stmt_line.memo = row[self.col_index['Mededeling']]
        stmt_line.amount = row[self.col_index['Bedrag']]
        stmt_line.payee = row[self.col_index['Naam tegenpartij']]
        stmt_line.refnum = row[self.col_index['Referentie']]
        
        stmt_line.trntype = self.dict_transaction_types[row[self.col_index['Beschrijving']]] or 'OTHER'
        if stmt_line.trntype == 'OTHER':
            logging.info('Other transaction type found: '+row[self.col_index['Beschrijving']])
        
        try:
            stmt_line.bank_account_to = BankAccountIban(row[self.col_index['Rekening tegenpartij']])
        except ValueError:
            pass
        
        return stmt_line

class BankAccountIban(BankAccount):
    """Represents a account id formatted in Belgian IBAN.
    Eg: BE46539007547034
    """    
    def __init__(self, acct_id):
        if acct_id and BankAccountIban.is_valid(acct_id):
            acct_id_canon = acct_id.replace(" ", "")
            super().__init__(bank_id=acct_id_canon[4:7], acct_id=acct_id_canon)
            self.acct_key = acct_id_canon[2:4]
        else:
            raise ValueError("acct_id is not a valid Belgian IBAN string.")
    
    iban_pattern = re.compile('^[A-Z]{2}\d{2} ?\d{4} ?\d{4} ?\d{4} ?[\d]{0,2}$')
    
    @staticmethod
    def is_valid(acct_id):
        """Returns True if the given account id is formatted as 
        for example 'BE46 5390 0754 7034' or 'BE46539007547034'
        """
        return BankAccountIban.iban_pattern.match(acct_id)

class ArgentaPlugin(Plugin):
    """Parses XLSX file from Belgian bank Argenta
    """
    def get_parser(self, fin):
        return ArgentaStatementParser(fin)